from datetime import datetime
import os
import pandas as pd
from openpyxl import load_workbook

class CsvProcessor:
    NUMBERS = "1234567890.-"

    def __init__(self, ):
        self.folder_path = os.path.abspath("Files/")

    def process_csv_to_excel(self, csv_file, excel_file, sheet_name):
        csv_df = pd.read_csv(os.path.join(self.folder_path, csv_file))
        workbook = load_workbook(os.path.join(self.folder_path, excel_file))
        sheet = workbook[sheet_name]

        for index, row in csv_df.iterrows():
            if "EN" in excel_file:
                row_name = row["NORMALS_ELEMENT"]
            else:
                row_name = row["ÉLÉMENT_DES_NORMALES"]
                
            row_index = self.find_row_index(sheet, row_name)

            if row_index != -1:
                self.update_excel_row(sheet, row, row_index)

        workbook.save(os.path.join(self.folder_path, excel_file))
        workbook.close()

    def find_row_index(self, sheet, row_name):
        for row_idx in range(4, 160):
            if row_name == sheet.cell(row=row_idx, column=19).value:
                return row_idx
        return -1

    def update_excel_row(self, sheet, row, row_index):
        df_index = 4
        for col_index in range(20, 34):
            val = row.iloc[df_index]
            if pd.notna(val):
                val = self.process_value(val)
                sheet.cell(row=row_index, column=col_index, value=val)
            df_index += 1

    def process_value(self, val):
        val = str(val).replace(',', '.')
        if all(char in self.NUMBERS for char in val):
            val = float(val)
        elif '/' in str(val):
            if ':' not in str(val):
                val = datetime.strptime(val, "%Y/%m/%d")
            else:
                val = datetime.strptime(val, "%Y/%m/%d %H:%M")
        return val

        