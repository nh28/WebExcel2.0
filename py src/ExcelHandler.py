from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
import re

class ExcelHandler:
    def __init__(self, excel_file_name, sheet_name, station_name):
        self.excel_file_path = "C:\\Users\\harschn\\Documents\\Climate Normal Verification\\Data Verification Live\\" + station_name + "\\" + excel_file_name
        try:
            self.workbook = load_workbook(filename=self.excel_file_path)
            self.sheet = self.workbook[sheet_name]
            self.NUMBERS = "1234567890.-,"
        except Exception as e:
            print("Invalid excel file: " + e)

    def write_excel(self, row_index, column_index, value):
        try:
            if value != "":
                if self._is_a_number(value):
                    if "," in value:
                        index = value.index(",")
                        value = value[:index] + "." + value[index + 1:]
                    if "." in value:
                        self.sheet.cell(row=row_index, column=column_index).value = float(value)
                    else:
                        self.sheet.cell(row=row_index, column=column_index).value = int(value)
                else:
                    self.sheet.cell(row=row_index, column=column_index).value = value

            self.workbook.save(filename=self.excel_file_path)

        except IndexError as e:
            print("Invalid row index:", e)
        except Exception as e:
            print(e)

    def find_index(self, target_string, row_max):
        try:
            for rowIndex in range(row_max):
                cell_value = self.sheet.cell(row=rowIndex + 1, column=1).value
                if cell_value is not None and target_string.lower() in cell_value.lower():
                    return rowIndex
            return -1
        except KeyError as e:
            print("Sheet does not exist:", e)
            return -1

    def _is_a_number(self, value):
        return all(char in self.NUMBERS for char in value)

    def find_value(self, row_index):
        try:
            return self.sheet.cell(row=row_index + 1, column=3).value
        except IndexError as e:
            print("Row index does not exist:", e)
            return None

    def close_workbook(self):
        try:
            self.workbook.close()
        except Exception as e:
            print(e)